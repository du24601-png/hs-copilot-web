"""Import the 2026 Chinese tariff legal notes workbook into hs_copilot.db.

The importer preserves the workbook text as authoritative source data and adds
only deterministic, machine-derived metadata for retrieval.  It never asks an
LLM to rewrite or interpret the source during import.
"""

import argparse
import hashlib
import json
import re
import sqlite3
from contextlib import closing
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


SOURCE_ID = "cn_tariff_2026"
DOCUMENT_TITLE = "中华人民共和国进出口税则（2026）"
REQUIRED_SHEETS = {"本国子目注释", "GRI归类总规则", "类注", "章注"}
STRICT_EXPECTED = {
    "legal_rule_count": 304,
    "decision_rule_count": 303,
    "compliance_notice_count": 1,
    "gri_count": 6,
    "section_count": 21,
    "section_note_count": 9,
    "chapter_count": 97,
    "chapter_note_count": 87,
    "national_scope_count": 243,
    "unmapped_national_scope_count": 0,
    "mapped_hs10_count": 340,
}

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS legal_source (
  source_id TEXT PRIMARY KEY,
  document_title TEXT NOT NULL,
  edition TEXT NOT NULL,
  jurisdiction TEXT NOT NULL,
  source_file TEXT NOT NULL,
  source_sha256 TEXT NOT NULL,
  imported_at TEXT NOT NULL,
  import_summary_json TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS tariff_scope (
  source_id TEXT NOT NULL,
  scope_type TEXT NOT NULL CHECK (scope_type IN ('section', 'chapter')),
  scope_ref TEXT NOT NULL,
  display_name TEXT NOT NULL,
  parent_scope_type TEXT,
  parent_scope_ref TEXT,
  note_status TEXT NOT NULL,
  source_sheet TEXT NOT NULL,
  source_row INTEGER NOT NULL,
  pdf_page TEXT,
  print_page TEXT,
  PRIMARY KEY (source_id, scope_type, scope_ref),
  FOREIGN KEY (source_id) REFERENCES legal_source(source_id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS legal_rule (
  rule_id TEXT PRIMARY KEY,
  source_id TEXT NOT NULL,
  rule_type TEXT NOT NULL CHECK (rule_type IN (
    'gri', 'section_note', 'chapter_note',
    'national_subheading_note', 'compliance_notice'
  )),
  ordinal INTEGER NOT NULL,
  title TEXT NOT NULL,
  full_text TEXT NOT NULL,
  status TEXT NOT NULL DEFAULT 'active',
  decision_eligible INTEGER NOT NULL CHECK (decision_eligible IN (0, 1)),
  source_sheet TEXT NOT NULL,
  source_row INTEGER NOT NULL,
  pdf_page TEXT,
  print_page TEXT,
  content_hash TEXT NOT NULL,
  FOREIGN KEY (source_id) REFERENCES legal_source(source_id) ON DELETE CASCADE,
  UNIQUE (source_id, rule_type, ordinal)
);

CREATE TABLE IF NOT EXISTS legal_rule_scope (
  rule_id TEXT NOT NULL,
  scope_type TEXT NOT NULL CHECK (scope_type IN ('global', 'section', 'chapter', 'subheading8')),
  scope_ref TEXT NOT NULL,
  PRIMARY KEY (rule_id, scope_type, scope_ref),
  FOREIGN KEY (rule_id) REFERENCES legal_rule(rule_id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS legal_clause (
  clause_id TEXT PRIMARY KEY,
  rule_id TEXT NOT NULL,
  clause_order INTEGER NOT NULL,
  clause_label TEXT,
  clause_text TEXT NOT NULL,
  effect_type TEXT NOT NULL,
  decision_eligible INTEGER NOT NULL CHECK (decision_eligible IN (0, 1)),
  content_hash TEXT NOT NULL,
  FOREIGN KEY (rule_id) REFERENCES legal_rule(rule_id) ON DELETE CASCADE,
  UNIQUE (rule_id, clause_order)
);

CREATE TABLE IF NOT EXISTS legal_relation (
  clause_id TEXT NOT NULL,
  relation_type TEXT NOT NULL,
  target_type TEXT NOT NULL CHECK (target_type IN ('section', 'chapter', 'heading', 'subheading8')),
  target_ref TEXT NOT NULL,
  evidence TEXT NOT NULL,
  PRIMARY KEY (clause_id, relation_type, target_type, target_ref),
  FOREIGN KEY (clause_id) REFERENCES legal_clause(clause_id) ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_tariff_scope_parent
  ON tariff_scope(source_id, parent_scope_type, parent_scope_ref);
CREATE INDEX IF NOT EXISTS idx_legal_rule_type
  ON legal_rule(source_id, rule_type, decision_eligible);
CREATE INDEX IF NOT EXISTS idx_legal_rule_scope_lookup
  ON legal_rule_scope(scope_type, scope_ref, rule_id);
CREATE INDEX IF NOT EXISTS idx_legal_clause_rule
  ON legal_clause(rule_id, clause_order);
CREATE INDEX IF NOT EXISTS idx_legal_relation_target
  ON legal_relation(target_type, target_ref, relation_type);

CREATE VIRTUAL TABLE IF NOT EXISTS legal_clause_fts USING fts5(
  clause_id UNINDEXED,
  source_id UNINDEXED,
  rule_id UNINDEXED,
  title,
  clause_text,
  tokenize='trigram'
);
"""


def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def sha256_text(value):
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def sha256_file(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_code_refs(value):
    text = clean_text(value).replace(".", "").replace("．", "")
    return list(dict.fromkeys(re.findall(r"(?<!\d)\d{8}(?!\d)", text)))


def split_clauses(text, rule_type):
    """Split long section/chapter notes only at top-level Chinese numbering."""
    source = clean_text(text)
    if not source:
        return []
    if rule_type not in {"section_note", "chapter_note"}:
        return [{"label": None, "text": source}]

    matches = list(re.finditer(r"(?m)^([一二三四五六七八九十百]+)、\s*", source))
    if not matches:
        return [{"label": None, "text": source}]

    clauses = []
    prefix = source[: matches[0].start()].strip()
    if prefix:
        clauses.append({"label": None, "text": prefix})
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(source)
        clauses.append({"label": match.group(1), "text": source[match.start() : end].strip()})
    return clauses


def detect_effect(text):
    value = clean_text(text)
    if re.search(r"禁止(?:进口|出口)", value):
        return "compliance"
    if re.search(r"不包括|除外|不归入", value):
        return "excludes"
    if re.search(r"应(?:分别)?归入|则归入|优先于", value):
        return "redirects"
    if re.search(r"所称|是指|定义", value):
        return "defines"
    if re.search(r"包括|视为", value):
        return "includes"
    if re.search(r"归类", value):
        return "classifies"
    return "other"


_CN_DIGITS = {"零": 0, "一": 1, "二": 2, "三": 3, "四": 4, "五": 5,
              "六": 6, "七": 7, "八": 8, "九": 9}


def chinese_number(value):
    if value.isdigit():
        return int(value)
    if value == "十":
        return 10
    if "十" in value:
        left, right = value.split("十", 1)
        return (_CN_DIGITS.get(left, 1) * 10) + _CN_DIGITS.get(right, 0)
    return _CN_DIGITS.get(value)


def extract_relations(text):
    value = clean_text(text).replace("．", ".")
    effect = detect_effect(value)
    relation_type = effect if effect in {"excludes", "redirects"} else "references"
    found = []

    for match in re.finditer(r"(?:本国子目|子目)\s*(\d{4})\s*[.]\s*(\d{4})", value):
        found.append(("subheading8", match.group(1) + match.group(2), match.group(0)))
    for match in re.finditer(r"(?:税目|品目)\s*(\d{2})\s*[.]\s*(\d{2})", value):
        found.append(("heading", match.group(1) + match.group(2), match.group(0)))
    for match in re.finditer(r"第\s*([一二三四五六七八九十百\d]{1,3})\s*章", value):
        number = chinese_number(match.group(1))
        if number is not None and 1 <= number <= 99:
            found.append(("chapter", f"{number:02d}", match.group(0)))
    for match in re.finditer(r"第\s*([一二三四五六七八九十百\d]{1,3})\s*类", value):
        number = chinese_number(match.group(1))
        if number is not None and 1 <= number <= 99:
            found.append(("section", f"{number:02d}", match.group(0)))

    relations = []
    seen = set()
    for target_type, target_ref, evidence in found:
        key = (relation_type, target_type, target_ref)
        if key in seen:
            continue
        seen.add(key)
        relations.append({
            "relation_type": relation_type,
            "target_type": target_type,
            "target_ref": target_ref,
            "evidence": evidence,
        })
    return relations


def _rows(sheet):
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(iterator)]
    for source_row, values in enumerate(iterator, start=2):
        yield source_row, dict(zip(headers, values))


def _page(value):
    text = clean_text(value)
    return text or None


def parse_workbook(path):
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    missing = REQUIRED_SHEETS.difference(workbook.sheetnames)
    if missing:
        raise ValueError("工作簿缺少工作表：" + "、".join(sorted(missing)))

    rules = []
    sections = []
    chapters = []

    for source_row, row in _rows(workbook["本国子目注释"]):
        text = clean_text(row.get("本国子目注释（完整原文）"))
        if not text:
            continue
        scope_refs = parse_code_refs(row.get("涉及税号（8位）"))
        if not scope_refs:
            scope_refs = parse_code_refs(row.get("涉及税号（点号）"))
        subject = clean_text(row.get("定义/对象"))
        is_notice = not subject and bool(re.search(r"禁止(?:进口|出口)", text))
        rule_type = "compliance_notice" if is_notice else "national_subheading_note"
        ordinal = source_row - 1
        rules.append({
            "rule_id": f"{SOURCE_ID}:{'notice' if is_notice else 'national'}:{ordinal:03d}",
            "rule_type": rule_type,
            "ordinal": ordinal,
            "title": (f"合规提示：{scope_refs[0]}" if is_notice else f"本国子目注释：{subject or scope_refs[0]}"),
            "full_text": text,
            "decision_eligible": 0 if is_notice else 1,
            "source_sheet": "本国子目注释",
            "source_row": source_row,
            "pdf_page": None,
            "print_page": None,
            "scope_type": "subheading8",
            "scope_refs": scope_refs,
        })

    for source_row, row in _rows(workbook["GRI归类总规则"]):
        text = clean_text(row.get("规则原文"))
        if not text:
            continue
        ordinal = int(row.get("序号"))
        rules.append({
            "rule_id": f"{SOURCE_ID}:gri:{ordinal:02d}",
            "rule_type": "gri",
            "ordinal": ordinal,
            "title": clean_text(row.get("规则")) or f"规则{ordinal}",
            "full_text": text,
            "decision_eligible": 1,
            "source_sheet": "GRI归类总规则",
            "source_row": source_row,
            "pdf_page": _page(row.get("PDF页码")),
            "print_page": _page(row.get("税则印刷页码")),
            "scope_type": "global",
            "scope_refs": ["*"],
        })

    for source_row, row in _rows(workbook["类注"]):
        ordinal = int(row.get("类序号"))
        section_ref = f"{ordinal:02d}"
        contained = clean_text(row.get("包含章节"))
        endpoints = [int(value) for value in re.findall(r"第\s*(\d+)\s*章", contained)]
        if not endpoints:
            raise ValueError(f"类注第 {source_row} 行无法解析包含章节：{contained}")
        chapter_start = endpoints[0]
        chapter_end = endpoints[-1]
        sections.append({
            "scope_ref": section_ref,
            "display_name": f"{clean_text(row.get('类'))} {clean_text(row.get('类名称'))}".strip(),
            "chapter_start": chapter_start,
            "chapter_end": chapter_end,
            "note_status": clean_text(row.get("状态")),
            "source_row": source_row,
            "pdf_page": _page(row.get("PDF页码")),
            "print_page": _page(row.get("税则印刷页码")),
        })
        text = clean_text(row.get("类注全文"))
        if text:
            rules.append({
                "rule_id": f"{SOURCE_ID}:section:{ordinal:02d}",
                "rule_type": "section_note",
                "ordinal": ordinal,
                "title": f"{clean_text(row.get('类'))}类注",
                "full_text": text,
                "decision_eligible": 1,
                "source_sheet": "类注",
                "source_row": source_row,
                "pdf_page": _page(row.get("PDF页码")),
                "print_page": _page(row.get("税则印刷页码")),
                "scope_type": "section",
                "scope_refs": [section_ref],
            })

    for source_row, row in _rows(workbook["章注"]):
        ordinal = int(row.get("章序号"))
        chapter_ref = f"{ordinal:02d}"
        section_number = re.search(r"(\d+)", clean_text(row.get("所属类")))
        section_ref = f"{int(section_number.group(1)):02d}" if section_number else None
        chapters.append({
            "scope_ref": chapter_ref,
            "display_name": f"{clean_text(row.get('章'))} {clean_text(row.get('章名称'))}".strip(),
            "parent_scope_ref": section_ref,
            "note_status": clean_text(row.get("状态")),
            "source_row": source_row,
            "pdf_page": _page(row.get("PDF页码")),
            "print_page": _page(row.get("税则印刷页码")),
        })
        text = clean_text(row.get("章注全文"))
        if text:
            rules.append({
                "rule_id": f"{SOURCE_ID}:chapter:{ordinal:02d}",
                "rule_type": "chapter_note",
                "ordinal": ordinal,
                "title": f"{clean_text(row.get('章'))}章注",
                "full_text": text,
                "decision_eligible": 1,
                "source_sheet": "章注",
                "source_row": source_row,
                "pdf_page": _page(row.get("PDF页码")),
                "print_page": _page(row.get("税则印刷页码")),
                "scope_type": "chapter",
                "scope_refs": [chapter_ref],
            })

    for rule in rules:
        clauses = []
        for order, clause in enumerate(split_clauses(rule["full_text"], rule["rule_type"]), start=1):
            clause_id = f"{rule['rule_id']}:c{order:02d}"
            clauses.append({
                "clause_id": clause_id,
                "clause_order": order,
                "clause_label": clause["label"],
                "clause_text": clause["text"],
                "effect_type": detect_effect(clause["text"]),
                "decision_eligible": rule["decision_eligible"],
                "content_hash": sha256_text(clause["text"]),
                "relations": extract_relations(clause["text"]),
            })
        rule["clauses"] = clauses
        rule["content_hash"] = sha256_text(rule["full_text"])

    workbook.close()
    return {"rules": rules, "sections": sections, "chapters": chapters}


def _table_exists(connection, name):
    return connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (name,)
    ).fetchone() is not None


def _fact_counts(connection):
    counts = {}
    for name in ("hs_code", "declare_element", "ciq_code", "hs_chapter"):
        if _table_exists(connection, name):
            counts[name] = connection.execute(f"SELECT COUNT(*) FROM {name}").fetchone()[0]
    return counts


def _summary(parsed, connection):
    rules = parsed["rules"]
    national_refs = {
        ref
        for rule in rules
        if rule["rule_type"] in {"national_subheading_note", "compliance_notice"}
        for ref in rule["scope_refs"]
    }
    mapped = set()
    mapped_scope_refs = set()
    if national_refs:
        placeholders = ",".join("?" for _ in national_refs)
        rows = connection.execute(
            f"SELECT code,subheading8 FROM hs_code WHERE subheading8 IN ({placeholders})",
            sorted(national_refs),
        ).fetchall()
        mapped = {row[0] for row in rows}
        mapped_scope_refs = {row[1] for row in rows}
    return {
        "legal_rule_count": len(rules),
        "decision_rule_count": sum(rule["decision_eligible"] for rule in rules),
        "compliance_notice_count": sum(rule["rule_type"] == "compliance_notice" for rule in rules),
        "gri_count": sum(rule["rule_type"] == "gri" for rule in rules),
        "section_count": len(parsed["sections"]),
        "section_note_count": sum(rule["rule_type"] == "section_note" for rule in rules),
        "chapter_count": len(parsed["chapters"]),
        "chapter_note_count": sum(rule["rule_type"] == "chapter_note" for rule in rules),
        "national_scope_count": len(national_refs),
        "unmapped_national_scope_count": len(national_refs - mapped_scope_refs),
        "mapped_hs10_count": len(mapped),
        "clause_count": sum(len(rule["clauses"]) for rule in rules),
        "relation_count": sum(
            len(clause["relations"])
            for rule in rules
            for clause in rule["clauses"]
        ),
    }


def _validate_strict(summary):
    mismatches = []
    for key, expected in STRICT_EXPECTED.items():
        if summary.get(key) != expected:
            mismatches.append(f"{key}: expected {expected}, got {summary.get(key)}")
    if mismatches:
        raise ValueError("2026 工作簿严格校验失败：" + "; ".join(mismatches))


def backup_database(db_path, backup_dir=None):
    source_path = Path(db_path).resolve()
    if not source_path.is_file():
        raise FileNotFoundError(f"数据库不存在：{source_path}")
    target_dir = Path(backup_dir).resolve() if backup_dir else source_path.parent / "tools" / "backup"
    target_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = target_dir / f"{source_path.stem}_before_legal_upgrade_{stamp}.db"
    with closing(sqlite3.connect(str(source_path))) as source:
        with closing(sqlite3.connect(str(target))) as output:
            source.backup(output)
    return target


def _remove_empty_prototype(connection):
    if not _table_exists(connection, "legal_note"):
        return
    count = connection.execute("SELECT COUNT(*) FROM legal_note").fetchone()[0]
    if count:
        raise RuntimeError(
            "发现非空 legal_note 原型表，已停止升级以避免丢失数据；请先迁移或人工确认。"
        )
    connection.execute("DROP TABLE IF EXISTS legal_note_fts")
    connection.execute("DROP TABLE legal_note")


def _execute_schema(connection):
    # sqlite3.executescript commits an open transaction implicitly.  Execute
    # each DDL statement ourselves so schema migration and data import remain
    # one rollback-safe transaction.
    for statement in SCHEMA_SQL.split(";"):
        sql = statement.strip()
        if sql:
            connection.execute(sql)


def import_legal_knowledge(xlsx_path, db_path, backup_enabled=True, strict=True, backup_dir=None):
    workbook_path = Path(xlsx_path).resolve()
    database_path = Path(db_path).resolve()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Excel 不存在：{workbook_path}")
    if not database_path.is_file():
        raise FileNotFoundError(f"数据库不存在：{database_path}")

    parsed = parse_workbook(workbook_path)
    with closing(sqlite3.connect(str(database_path))) as validation_connection:
        if not _table_exists(validation_connection, "hs_code"):
            raise RuntimeError("目标数据库缺少 hs_code 表")
        columns = {row[1] for row in validation_connection.execute("PRAGMA table_info(hs_code)")}
        if not {"code", "subheading8", "chapter"}.issubset(columns):
            raise RuntimeError("hs_code 缺少 code/subheading8/chapter 字段")
        summary = _summary(parsed, validation_connection)
        facts_before = _fact_counts(validation_connection)
    if strict:
        _validate_strict(summary)

    backup_path = backup_database(database_path, backup_dir) if backup_enabled else None
    imported_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    source_hash = sha256_file(workbook_path)

    connection = sqlite3.connect(str(database_path))
    connection.execute("PRAGMA foreign_keys=ON")
    try:
        connection.execute("BEGIN IMMEDIATE")
        _remove_empty_prototype(connection)
        _execute_schema(connection)

        connection.execute("DELETE FROM legal_clause_fts WHERE source_id=?", (SOURCE_ID,))
        connection.execute("DELETE FROM legal_source WHERE source_id=?", (SOURCE_ID,))
        connection.execute(
            """INSERT INTO legal_source(
                 source_id, document_title, edition, jurisdiction, source_file,
                 source_sha256, imported_at, import_summary_json
               ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                SOURCE_ID,
                DOCUMENT_TITLE,
                "2026",
                "CN",
                workbook_path.name,
                source_hash,
                imported_at,
                json.dumps(summary, ensure_ascii=False, sort_keys=True),
            ),
        )

        for section in parsed["sections"]:
            connection.execute(
                """INSERT INTO tariff_scope VALUES (?, 'section', ?, ?, NULL, NULL, ?, '类注', ?, ?, ?)""",
                (
                    SOURCE_ID,
                    section["scope_ref"],
                    section["display_name"],
                    section["note_status"],
                    section["source_row"],
                    section["pdf_page"],
                    section["print_page"],
                ),
            )
        for chapter in parsed["chapters"]:
            connection.execute(
                """INSERT INTO tariff_scope VALUES (?, 'chapter', ?, ?, 'section', ?, ?, '章注', ?, ?, ?)""",
                (
                    SOURCE_ID,
                    chapter["scope_ref"],
                    chapter["display_name"],
                    chapter["parent_scope_ref"],
                    chapter["note_status"],
                    chapter["source_row"],
                    chapter["pdf_page"],
                    chapter["print_page"],
                ),
            )

        for rule in parsed["rules"]:
            connection.execute(
                """INSERT INTO legal_rule(
                     rule_id, source_id, rule_type, ordinal, title, full_text,
                     status, decision_eligible, source_sheet, source_row,
                     pdf_page, print_page, content_hash
                   ) VALUES (?, ?, ?, ?, ?, ?, 'active', ?, ?, ?, ?, ?, ?)""",
                (
                    rule["rule_id"], SOURCE_ID, rule["rule_type"], rule["ordinal"],
                    rule["title"], rule["full_text"], rule["decision_eligible"],
                    rule["source_sheet"], rule["source_row"], rule["pdf_page"],
                    rule["print_page"], rule["content_hash"],
                ),
            )
            for scope_ref in rule["scope_refs"]:
                connection.execute(
                    "INSERT INTO legal_rule_scope VALUES (?, ?, ?)",
                    (rule["rule_id"], rule["scope_type"], scope_ref),
                )
            for clause in rule["clauses"]:
                connection.execute(
                    """INSERT INTO legal_clause VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        clause["clause_id"], rule["rule_id"], clause["clause_order"],
                        clause["clause_label"], clause["clause_text"], clause["effect_type"],
                        clause["decision_eligible"], clause["content_hash"],
                    ),
                )
                connection.execute(
                    "INSERT INTO legal_clause_fts VALUES (?, ?, ?, ?, ?)",
                    (
                        clause["clause_id"], SOURCE_ID, rule["rule_id"],
                        rule["title"], clause["clause_text"],
                    ),
                )
                for relation in clause["relations"]:
                    connection.execute(
                        "INSERT INTO legal_relation VALUES (?, ?, ?, ?, ?)",
                        (
                            clause["clause_id"], relation["relation_type"],
                            relation["target_type"], relation["target_ref"],
                            relation["evidence"],
                        ),
                    )

        foreign_key_errors = connection.execute("PRAGMA foreign_key_check").fetchall()
        if foreign_key_errors:
            raise RuntimeError(f"外键检查失败：{foreign_key_errors[:3]}")
        facts_after = _fact_counts(connection)
        if facts_after != facts_before:
            raise RuntimeError(f"原有事实表计数发生变化：before={facts_before}, after={facts_after}")
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()

    result = dict(summary)
    result["source_id"] = SOURCE_ID
    result["source_sha256"] = source_hash
    result["backup_path"] = str(backup_path) if backup_path else None
    result["fact_counts"] = facts_before
    return result


def main():
    parser = argparse.ArgumentParser(description="导入 2026 税则法律规则知识层")
    parser.add_argument("--xlsx", required=True, help="规则 Excel 文件路径")
    parser.add_argument("--db", default="hs_copilot.db", help="目标 SQLite 数据库")
    parser.add_argument("--backup-dir", help="备份目录；默认 tools/backup")
    parser.add_argument("--no-backup", action="store_true", help="跳过备份，仅用于测试")
    parser.add_argument("--no-strict", action="store_true", help="跳过 2026 固定统计校验，仅用于 fixture")
    args = parser.parse_args()
    result = import_legal_knowledge(
        args.xlsx,
        args.db,
        backup_enabled=not args.no_backup,
        strict=not args.no_strict,
        backup_dir=args.backup_dir,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
