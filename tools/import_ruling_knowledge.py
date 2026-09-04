"""Import the approved 104 historical rulings without interpreting their text.

Historical codes are prefix-linked to the local 2026 catalog, never promoted to
confirmed current classifications. Missing dates, narratives and URLs stay null;
the source validity wording is not a present-day validity assertion.

Dry runs copy database bytes and existing WAL/journal into a temporary directory.
They never open the source SQLite database, including its shared-memory sidecar.
Formal imports always back up through SQLite before any schema/data change.
"""

import argparse
import hashlib
import io
import json
import re
import shutil
import sqlite3
import tempfile
import uuid
from contextlib import closing, contextmanager
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_ID = "cn_rulings_2023_2025"
TARIFF_YEAR = 2026
ORIGINAL_COLUMNS = (
    "year", "decision_no", "case_type", "hs_code_original", "heading_4",
    "subheading_6", "code_level", "product_name_cn", "product_name_en",
    "specification", "product_description", "classification_decision", "rule_basis",
    "announcement_no", "publish_date", "effective_date", "validity_status",
    "replaced_by", "verification_grade", "content_source", "source_file",
    "content_source_url", "official_notice_url", "raw_text_status",
    "current_code_2026", "code_status_2026", "knowledge_status",
)
STRICT_EXPECTED = {
    "ruling_case_count": 104,
    "approved_count": 104,
    "unique_prefix_count": 27,
    "multiple_prefix_count": 77,
    "missing_prefix_count": 0,
    "publish_date_missing_count": 93,
    "classification_decision_missing_count": 24,
    "validity_unconfirmed_count": 104,
}
OWNED_TABLES = {"ruling_source", "ruling_case", "ruling_code_map", "ruling_case_fts"} | {
    "ruling_case_fts_" + suffix for suffix in ("data", "idx", "content", "docsize", "config")
}

# Execute statements individually inside BEGIN IMMEDIATE. executescript would
# implicitly commit and break atomic rollback of DDL, replacement, and FTS.
SCHEMA_STATEMENTS = (
    """CREATE TABLE IF NOT EXISTS ruling_source (
        source_id TEXT PRIMARY KEY,
        source_file TEXT NOT NULL,
        source_sha256 TEXT NOT NULL,
        imported_at TEXT NOT NULL,
        import_summary_json TEXT NOT NULL
    )""",
    "CREATE TABLE IF NOT EXISTS ruling_case ("
    "case_id TEXT PRIMARY KEY, source_id TEXT NOT NULL, "
    + ", ".join('"' + column + '" ' + ("INTEGER" if column == "year" else "TEXT")
                for column in ORIGINAL_COLUMNS)
    + ", historical_code TEXT NOT NULL, source_row INTEGER NOT NULL, "
    "content_hash TEXT NOT NULL, original_json TEXT NOT NULL, "
    "FOREIGN KEY(source_id) REFERENCES ruling_source(source_id) ON DELETE CASCADE)",
    """CREATE TABLE IF NOT EXISTS ruling_code_map (
        case_id TEXT NOT NULL,
        code TEXT NOT NULL,
        tariff_year INTEGER NOT NULL,
        match_method TEXT NOT NULL DEFAULT 'prefix' CHECK(match_method='prefix'),
        review_status TEXT NOT NULL DEFAULT 'pending',
        PRIMARY KEY(case_id, code, tariff_year),
        FOREIGN KEY(case_id) REFERENCES ruling_case(case_id) ON DELETE CASCADE,
        FOREIGN KEY(code) REFERENCES hs_code(code)
    )""",
    "CREATE INDEX IF NOT EXISTS idx_ruling_historical_code ON ruling_case(historical_code)",
    "CREATE INDEX IF NOT EXISTS idx_ruling_heading4 ON ruling_case(heading_4)",
    "CREATE INDEX IF NOT EXISTS idx_ruling_subheading6 ON ruling_case(subheading_6)",
    "CREATE INDEX IF NOT EXISTS idx_ruling_source ON ruling_case(source_id)",
    "CREATE INDEX IF NOT EXISTS idx_ruling_map_code ON ruling_code_map(code, tariff_year)",
    """CREATE VIRTUAL TABLE IF NOT EXISTS ruling_case_fts USING fts5(
        case_id UNINDEXED, source_id UNINDEXED, product_name,
        product_description, classification_decision, rule_basis, tokenize='trigram'
    )""",
)


def _json(value):
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), allow_nan=False)


def _sha256_file(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalize_code(value, column, length=None):
    if not isinstance(value, str) or not re.fullmatch(r"[0-9.．\s]+", value):
        raise ValueError(f"{column}: expected a text tariff code, got {value!r}")
    normalized = re.sub(r"[.．\s]", "", value)
    allowed_lengths = (length,) if length else (6, 8)
    if len(normalized) not in allowed_lengths:
        raise ValueError(f"{column}: invalid code length {len(normalized)}")
    return normalized


def parse_workbook(xlsx_path):
    """Read a single byte snapshot so file provenance and parsed values agree."""
    source_bytes = Path(xlsx_path).read_bytes()
    cases = []
    seen = set()
    workbook = load_workbook(io.BytesIO(source_bytes), read_only=True, data_only=False)
    try:
        if "判例知识库" not in workbook.sheetnames:
            raise ValueError("工作簿缺少判例知识库工作表")
        iterator = workbook["判例知识库"].iter_rows()
        header = tuple(cell.value for cell in next(iterator, ()))
        if header != ORIGINAL_COLUMNS:
            raise ValueError("判例知识库表头必须包含原始 27 个字段，且顺序不变")
        for source_row, cells in enumerate(iterator, start=2):
            values = [cell.value for cell in cells]
            if all(value is None for value in values):
                continue
            if len(values) != len(ORIGINAL_COLUMNS):
                raise ValueError(f"row {source_row}: expected 27 columns")
            if any(cell.data_type == "f" for cell in cells):
                raise ValueError(f"row {source_row}: formulas are not original source values")
            original = dict(zip(ORIGINAL_COLUMNS, values))
            for column, value in original.items():
                if column == "year":
                    if type(value) is not int or not 1900 <= value <= 9999:
                        raise ValueError(f"row {source_row} year: expected integer year")
                elif value is not None and not isinstance(value, str):
                    raise ValueError(f"row {source_row} {column}: expected original text or null")
            decision_no = original["decision_no"]
            if not decision_no or decision_no != decision_no.strip() or decision_no in seen:
                raise ValueError(f"row {source_row} decision_no: blank, padded or duplicate identifier")
            seen.add(decision_no)
            if original["knowledge_status"] != "可入库":
                raise ValueError(f"row {source_row} knowledge_status: only 可入库 is approved")
            code = _normalize_code(original["hs_code_original"], "hs_code_original")
            if original["code_level"] != f"{len(code)}位":
                raise ValueError(f"row {source_row} code_level disagrees with hs_code_original")
            for column, length in (("heading_4", 4), ("subheading_6", 6)):
                if _normalize_code(original[column], column, length) != code[:length]:
                    raise ValueError(f"row {source_row} {column} disagrees with hs_code_original")
            raw_json = _json(original)
            cases.append({
                "case_id": "cn_ruling:" + decision_no,
                "original": original,
                "historical_code": code,
                "source_row": source_row,
                "original_json": raw_json,
                "content_hash": hashlib.sha256(raw_json.encode("utf-8")).hexdigest(),
            })
    finally:
        workbook.close()
    return {"cases": cases, "source_sha256": hashlib.sha256(source_bytes).hexdigest()}


def _database_file_hashes(db_path):
    # SHM is derived lock/index state, never copied or connected to on the source.
    return {
        suffix: _sha256_file(path)
        for suffix in ("", "-wal", "-journal")
        if (path := Path(str(db_path) + suffix)).is_file()
    }


@contextmanager
def _read_only_snapshot(db_path):
    """Refuse a changing source instead of returning an inconsistent dry run."""
    if not db_path.is_file():
        raise FileNotFoundError(f"数据库不存在：{db_path}")
    before = _database_file_hashes(db_path)
    with tempfile.TemporaryDirectory(prefix="ruling-preflight-") as temp:
        snapshot = Path(temp) / "snapshot.db"
        try:
            for suffix in before:
                shutil.copyfile(Path(str(db_path) + suffix), Path(str(snapshot) + suffix))
            copied = _database_file_hashes(snapshot)
            after = _database_file_hashes(db_path)
        except FileNotFoundError as error:
            raise RuntimeError("数据库快照复制期间源文件变化；请停止写入后重试") from error
        if before != copied or before != after:
            raise RuntimeError("数据库快照复制期间内容发生变化 (source changed)；请停止写入后重试")
        # Only the disposable copy may recover a journal or build WAL/SHM files.
        with closing(sqlite3.connect(snapshot)) as connection:
            connection.execute("PRAGMA query_only=ON")
            connection.execute("BEGIN")
            yield connection
        if _database_file_hashes(db_path) != before:
            raise RuntimeError("数据库预检期间内容发生变化 (source changed)；请停止写入后重试")


def _table_exists(connection, table):
    return connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)
    ).fetchone() is not None


def _content_fingerprint(rows):
    # Row hashes are sorted rather than relying on rowid, primary-key shape, or
    # collation. The type tag distinguishes BLOB bytes from their textual form.
    hashes = []
    for row in rows:
        encoded = _json([
            ["blob", value.hex()] if isinstance(value, bytes) else [type(value).__name__, value]
            for value in row
        ])
        hashes.append(hashlib.sha256(encoded.encode("utf-8")).digest())
    return {"count": len(hashes), "sha256": hashlib.sha256(b"".join(sorted(hashes))).hexdigest()}


def non_ruling_fingerprints(connection):
    """Count and hash every original non-ruling table, including FTS shadows."""
    names = [row[0] for row in connection.execute(
        "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
    ) if row[0] not in OWNED_TABLES]
    return {
        name: _content_fingerprint(connection.execute('SELECT * FROM "' + name.replace('"', '""') + '"'))
        for name in names
    }


def _other_source_fingerprints(connection):
    if not _table_exists(connection, "ruling_source"):
        return {}
    queries = {
        "ruling_source": "SELECT * FROM ruling_source WHERE source_id<>?",
        "ruling_case": "SELECT * FROM ruling_case WHERE source_id<>?",
        "ruling_code_map": "SELECT m.* FROM ruling_code_map m JOIN ruling_case c USING(case_id) WHERE c.source_id<>?",
        "ruling_case_fts": "SELECT * FROM ruling_case_fts WHERE source_id<>?",
    }
    return {name: _content_fingerprint(connection.execute(query, (SOURCE_ID,)))
            for name, query in queries.items()}


def _map_and_summarize(parsed, connection, strict):
    catalog = [row[0] for row in connection.execute("SELECT code FROM hs_code ORDER BY code")]
    if any(not isinstance(code, str) or not re.fullmatch(r"[0-9]{10}", code) for code in catalog):
        raise ValueError("hs_code: local catalog must contain 10-digit text codes")
    mappings = {
        case["case_id"]: [code for code in catalog if code.startswith(case["historical_code"])]
        for case in parsed["cases"]
    }
    originals = [case["original"] for case in parsed["cases"]]
    summary = {
        "ruling_case_count": len(originals),
        "approved_count": sum(row["knowledge_status"] == "可入库" for row in originals),
        "unique_prefix_count": sum(len(codes) == 1 for codes in mappings.values()),
        "multiple_prefix_count": sum(len(codes) > 1 for codes in mappings.values()),
        "missing_prefix_count": sum(not codes for codes in mappings.values()),
        "mapping_count": sum(len(codes) for codes in mappings.values()),
        "publish_date_missing_count": sum(row["publish_date"] is None for row in originals),
        "classification_decision_missing_count": sum(row["classification_decision"] is None for row in originals),
        "validity_unconfirmed_count": sum(row["validity_status"] == "未在所附2025失效清单中" for row in originals),
        "tariff_year": TARIFF_YEAR,
        "match_method": "prefix",
        "review_status": "pending",
    }
    if strict:
        errors = [f"{key}: expected {expected}, got {summary[key]}"
                  for key, expected in STRICT_EXPECTED.items() if summary[key] != expected]
        if errors:
            raise ValueError("判例工作簿严格校验失败 (strict): " + "; ".join(errors))
    return mappings, summary


def _backup_database(db_path, backup_dir, expected_fingerprints):
    destination_dir = Path(backup_dir).resolve() if backup_dir else PROJECT_ROOT / "tools" / "backup"
    destination_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S.%fZ")
    destination = destination_dir / f"{db_path.stem}-before-rulings-{stamp}-{uuid.uuid4().hex[:8]}.db"
    # The main importer already holds BEGIN IMMEDIATE: other writers cannot race
    # this separate read connection, and SQLite backup includes committed WAL.
    try:
        with closing(sqlite3.connect(db_path.as_uri() + "?mode=ro", uri=True)) as source:
            with closing(sqlite3.connect(destination)) as target:
                source.backup(target)
                if target.execute("PRAGMA integrity_check").fetchone()[0] != "ok":
                    raise RuntimeError("备份完整性检查失败")
                if non_ruling_fingerprints(target) != expected_fingerprints:
                    raise RuntimeError("备份原有表内容与导入前快照不一致")
    except Exception:
        # This exact newly-created path belongs to this operation, never a glob.
        destination.unlink(missing_ok=True)
        raise
    return destination


def _replace_source(connection, parsed, source_path, summary):
    for statement in SCHEMA_STATEMENTS:
        connection.execute(statement)
    connection.execute("DELETE FROM ruling_case_fts WHERE source_id=?", (SOURCE_ID,))
    connection.execute("DELETE FROM ruling_source WHERE source_id=?", (SOURCE_ID,))
    connection.execute("INSERT INTO ruling_source VALUES (?, ?, ?, ?, ?)", (
        SOURCE_ID, source_path.name, parsed["source_sha256"],
        datetime.now(timezone.utc).isoformat(), _json(summary),
    ))
    columns = ("case_id", "source_id") + ORIGINAL_COLUMNS + (
        "historical_code", "source_row", "content_hash", "original_json",
    )
    insert_case = "INSERT INTO ruling_case (" + ",".join(columns) + ") VALUES (" + ",".join("?" for _ in columns) + ")"
    for case in parsed["cases"]:
        original = case["original"]
        connection.execute(insert_case, (
            case["case_id"], SOURCE_ID, *(original[column] for column in ORIGINAL_COLUMNS),
            case["historical_code"], case["source_row"], case["content_hash"], case["original_json"],
        ))
        names = [original[column] for column in ("product_name_cn", "product_name_en") if original[column]]
        connection.execute("""INSERT INTO ruling_case_fts
            (case_id,source_id,product_name,product_description,classification_decision,rule_basis)
            VALUES (?, ?, ?, ?, ?, ?)""", (
            case["case_id"], SOURCE_ID, "\n".join(names), original["product_description"],
            original["classification_decision"], original["rule_basis"],
        ))


def import_ruling_knowledge(xlsx_path, db_path=None, *, dry_run=False, strict=True, backup_dir=None):
    """Validate then atomically replace this source; strict=False is for fixtures.

    There is deliberately no no-backup switch. Validation always precedes backup
    creation, and dry_run only writes disposable files in the system temp dir.
    """
    source_path = Path(xlsx_path).resolve()
    database_path = Path(db_path).resolve() if db_path else PROJECT_ROOT / "hs_copilot.db"
    if source_path == database_path:
        raise ValueError("Source workbook and destination database must be different files")
    parsed = parse_workbook(source_path)
    with _read_only_snapshot(database_path) as snapshot:
        _, summary = _map_and_summarize(parsed, snapshot, strict)
        before = non_ruling_fingerprints(snapshot)
    result = dict(summary, source_id=SOURCE_ID, source_sha256=parsed["source_sha256"],
                  dry_run=bool(dry_run), backup_path=None,
                  non_ruling_before=before, non_ruling_after=before)
    if dry_run:
        return result

    with closing(sqlite3.connect(database_path.as_uri() + "?mode=rw", uri=True)) as connection:
        connection.execute("PRAGMA foreign_keys=ON")
        connection.execute("PRAGMA busy_timeout=5000")
        try:
            connection.execute("BEGIN IMMEDIATE")
            # Re-evaluate under the writer lock. No stale preflight mappings can
            # slip into the import if the catalog changed between the two phases.
            mappings, summary = _map_and_summarize(parsed, connection, strict)
            before = non_ruling_fingerprints(connection)
            other_before = _other_source_fingerprints(connection)
            backup_path = _backup_database(database_path, backup_dir, before)
            _replace_source(connection, parsed, source_path, summary)
            for case_id, codes in mappings.items():
                connection.executemany("INSERT INTO ruling_code_map VALUES (?, ?, ?, 'prefix', 'pending')",
                                       ((case_id, code, TARIFF_YEAR) for code in codes))
            foreign_key_errors = connection.execute("PRAGMA foreign_key_check").fetchall()
            if foreign_key_errors:
                raise RuntimeError(f"外键检查失败：{foreign_key_errors[:3]}")
            if connection.execute("PRAGMA integrity_check").fetchone()[0] != "ok":
                raise RuntimeError("导入后数据库完整性检查失败")
            after = non_ruling_fingerprints(connection)
            if before != after:
                raise RuntimeError("原有 non-ruling 表行数或内容哈希发生变化；已阻止提交")
            other_after = _other_source_fingerprints(connection)
            if other_before and other_before != other_after:
                raise RuntimeError("其他判例来源发生变化；已阻止提交")
            connection.commit()
        except Exception:
            connection.rollback()
            raise
    return dict(summary, source_id=SOURCE_ID, source_sha256=parsed["source_sha256"],
                dry_run=False, backup_path=str(backup_path),
                non_ruling_before=before, non_ruling_after=after)


def main():
    parser = argparse.ArgumentParser(description="导入经批准的历史判例知识库；前缀映射均待人工确认")
    parser.add_argument("--xlsx", required=True, help="原始判例 Excel 工作簿")
    parser.add_argument("--db", default=str(PROJECT_ROOT / "hs_copilot.db"), help="目标 SQLite 数据库")
    parser.add_argument("--dry-run", action="store_true", help="只读预检，不写数据库或其 WAL/SHM")
    parser.add_argument("--backup-dir", help="正式导入备份目录，默认 tools/backup")
    args = parser.parse_args()
    result = import_ruling_knowledge(args.xlsx, args.db, dry_run=args.dry_run, backup_dir=args.backup_dir)
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
