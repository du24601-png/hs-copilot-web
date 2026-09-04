"""Importer regression tests; every database mutation is confined to a temp dir."""

import hashlib
import importlib.util
import json
import shutil
import sqlite3
import tempfile
import unittest
from contextlib import closing
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
HEADERS = [
    "year", "decision_no", "case_type", "hs_code_original", "heading_4",
    "subheading_6", "code_level", "product_name_cn", "product_name_en",
    "specification", "product_description", "classification_decision", "rule_basis",
    "announcement_no", "publish_date", "effective_date", "validity_status",
    "replaced_by", "verification_grade", "content_source", "source_file",
    "content_source_url", "official_notice_url", "raw_text_status",
    "current_code_2026", "code_status_2026", "knowledge_status",
]
MODULE_PATH = ROOT / "tools" / "import_ruling_knowledge.py"
if MODULE_PATH.exists():
    spec = importlib.util.spec_from_file_location("ruling_importer", MODULE_PATH)
    importer = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(importer)
else:
    importer = None


def database_files(db):
    return {
        suffix: hashlib.sha256(path.read_bytes()).hexdigest()
        for suffix in ("", "-wal", "-shm", "-journal")
        if (path := Path(str(db) + suffix)).exists()
    }


class RulingKnowledgeImportTest(unittest.TestCase):
    def setUp(self):
        self.assertIsNotNone(importer, "ruling importer has not been implemented")
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.root = Path(self.tmp.name)
        self.xlsx = self.root / "fixture.xlsx"
        self.db = self.root / "fixture.db"
        self.backups = self.root / "backups"
        common = dict.fromkeys(HEADERS)
        common.update(
            year=2023, case_type="J类归类决定", product_name_cn="测试平衡车",
            product_description=" 原始描述\n保留空白。 ",
            classification_decision="根据原始依据归入。", rule_basis="归类总规则一及六",
            announcement_no="海关总署公告2023年第94号",
            validity_status="未在所附2025失效清单中", verification_grade="A",
            content_source="用户提供的官方公告附件", source_file="原文.doc",
            official_notice_url="https://customs.gov.cn/.../notice", raw_text_status="官方附件原文",
            current_code_2026="9999999999", code_status_2026="故意错误的源提示",
            knowledge_status="可入库",
        )
        self.rows = [
            dict(common, decision_no="J2023-0001", hs_code_original="8711.6000",
                 heading_4="8711", subheading_6="871160", code_level="8位"),
            dict(common, decision_no="W2023-1", case_type="WCO转化决定(W)",
                 hs_code_original="1704.90", heading_4="1704", subheading_6="170490",
                 code_level="6位", product_name_cn="草本止咳糖", classification_decision=None),
        ]
        self.write_workbook()
        with closing(sqlite3.connect(self.db)) as connection:
            connection.executescript("""
                CREATE TABLE hs_code(code TEXT PRIMARY KEY, name TEXT, subheading8 TEXT);
                INSERT INTO hs_code VALUES('8711600000','电动车','87116000');
                INSERT INTO hs_code VALUES('1704900001','糖一','17049000');
                INSERT INTO hs_code VALUES('1704900002','糖二','17049000');
                CREATE TABLE legal_rule(rule_id TEXT PRIMARY KEY, full_text TEXT, payload BLOB);
                INSERT INTO legal_rule VALUES('rule:1','不可修改的规则',X'00FF');
                CREATE VIRTUAL TABLE legal_clause_fts USING fts5(clause_text, tokenize='trigram');
                INSERT INTO legal_clause_fts VALUES('原有法律规则');
            """)
            connection.commit()

    def write_workbook(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "判例知识库"
        sheet.append(HEADERS)
        for row in self.rows:
            sheet.append([row.get(column) for column in HEADERS])
        workbook.save(self.xlsx)
        workbook.close()

    def run_import(self, **kwargs):
        return importer.import_ruling_knowledge(
            self.xlsx, self.db, backup_dir=self.backups, strict=False, **kwargs
        )

    def dump_database(self):
        with closing(sqlite3.connect(self.db)) as connection:
            return list(connection.iterdump())

    def test_dry_run_leaves_database_bytes_and_sidecars_unchanged(self):
        before = database_files(self.db)
        result = self.run_import(dry_run=True)
        self.assertEqual(database_files(self.db), before)
        self.assertFalse(self.backups.exists())
        self.assertEqual(result["ruling_case_count"], 2)
        self.assertEqual(result["unique_prefix_count"], 1)
        self.assertEqual(result["multiple_prefix_count"], 1)
        self.assertEqual(result["missing_prefix_count"], 0)
        self.assertEqual(result["classification_decision_missing_count"], 1)
        self.assertEqual(result["publish_date_missing_count"], 2)
        self.assertTrue(result["dry_run"])

    def test_dry_run_reads_committed_wal_without_touching_source_sidecars(self):
        with closing(sqlite3.connect(self.db)) as writer:
            writer.execute("PRAGMA journal_mode=WAL")
            writer.execute("INSERT INTO hs_code VALUES('8711600001','新候选','87116000')")
            writer.commit()
            before = database_files(self.db)
            self.assertIn("-wal", before)
            result = self.run_import(dry_run=True)
            self.assertEqual(result["multiple_prefix_count"], 2)
            self.assertEqual(database_files(self.db), before)

    def test_snapshot_refuses_source_changed_during_copy(self):
        copyfile = shutil.copyfile

        def concurrent_write(src, dst, *args, **kwargs):
            copied = copyfile(src, dst, *args, **kwargs)
            if Path(src).resolve() == self.db.resolve():
                with closing(sqlite3.connect(self.db)) as writer:
                    writer.execute("INSERT INTO hs_code VALUES('8711600001','新候选','87116000')")
                    writer.commit()
            return copied

        with patch.object(importer.shutil, "copyfile", side_effect=concurrent_write):
            with self.assertRaisesRegex(RuntimeError, "changed|变化|写入"):
                self.run_import(dry_run=True)
        self.assertFalse(self.backups.exists())

    def test_preserves_all_original_fields_nulls_historical_code_and_hashes(self):
        result = self.run_import()
        self.assertEqual(result["source_sha256"], hashlib.sha256(self.xlsx.read_bytes()).hexdigest())
        with closing(sqlite3.connect(self.db)) as connection:
            connection.row_factory = sqlite3.Row
            rows = connection.execute("SELECT * FROM ruling_case ORDER BY source_row").fetchall()
            for source_row, (expected, stored) in enumerate(zip(self.rows, rows), start=2):
                self.assertEqual(json.loads(stored["original_json"]), expected)
                self.assertEqual({column: stored[column] for column in HEADERS}, expected)
                self.assertEqual(stored["source_row"], source_row)
                self.assertEqual(stored["case_id"], "cn_ruling:" + expected["decision_no"])
                self.assertEqual(stored["content_hash"], hashlib.sha256(stored["original_json"].encode()).hexdigest())
            self.assertEqual(rows[0]["historical_code"], "87116000")
            self.assertEqual(rows[1]["historical_code"], "170490")
            self.assertIsNone(rows[1]["classification_decision"])

    def test_recomputes_prefix_mapping_and_keeps_every_candidate_pending(self):
        self.run_import()
        with closing(sqlite3.connect(self.db)) as connection:
            mappings = connection.execute(
                "SELECT case_id,code,tariff_year,match_method,review_status FROM ruling_code_map ORDER BY case_id,code"
            ).fetchall()
            self.assertEqual(mappings, [
                ("cn_ruling:J2023-0001", "8711600000", 2026, "prefix", "pending"),
                ("cn_ruling:W2023-1", "1704900001", 2026, "prefix", "pending"),
                ("cn_ruling:W2023-1", "1704900002", 2026, "prefix", "pending"),
            ])
            self.assertEqual(connection.execute("SELECT count(*) FROM ruling_case_fts").fetchone()[0], 2)
            self.assertEqual(connection.execute(
                "SELECT case_id FROM ruling_case_fts WHERE ruling_case_fts MATCH ?", ('草本止咳糖',)
            ).fetchone()[0], "cn_ruling:W2023-1")
            self.assertEqual(connection.execute("PRAGMA foreign_key_check").fetchall(), [])

    def test_repeat_import_is_idempotent_and_backups_are_distinct_preimages(self):
        first = self.run_import()
        first_backup = Path(first["backup_path"])
        second = self.run_import()
        self.assertNotEqual(first["backup_path"], second["backup_path"])
        with closing(sqlite3.connect(first_backup)) as connection:
            self.assertIsNone(connection.execute("SELECT name FROM sqlite_master WHERE name='ruling_case'").fetchone())
        with closing(sqlite3.connect(second["backup_path"])) as connection:
            self.assertEqual(connection.execute("SELECT count(*) FROM ruling_case").fetchone()[0], 2)
            self.assertEqual(connection.execute("PRAGMA integrity_check").fetchone()[0], "ok")
        with closing(sqlite3.connect(self.db)) as connection:
            self.assertEqual(connection.execute("SELECT count(*) FROM ruling_source").fetchone()[0], 1)
            self.assertEqual(connection.execute("SELECT count(*) FROM ruling_case").fetchone()[0], 2)
            self.assertEqual(connection.execute("SELECT count(*) FROM ruling_code_map").fetchone()[0], 3)

    def test_atomic_replacement_preserves_other_sources(self):
        self.run_import()
        with closing(sqlite3.connect(self.db)) as connection:
            connection.execute("INSERT INTO ruling_source SELECT 'other',source_file,source_sha256,imported_at,import_summary_json FROM ruling_source")
            columns = [row[1] for row in connection.execute("PRAGMA table_info(ruling_case)")]
            row = list(connection.execute("SELECT * FROM ruling_case LIMIT 1").fetchone())
            row[columns.index("case_id")] = "other:case"
            row[columns.index("source_id")] = "other"
            connection.execute("INSERT INTO ruling_case VALUES (" + ",".join("?" for _ in row) + ")", row)
            connection.execute("INSERT INTO ruling_code_map VALUES ('other:case','8711600000',2026,'prefix','pending')")
            connection.execute("INSERT INTO ruling_case_fts VALUES ('other:case','other','其他原文',NULL,NULL,NULL)")
            connection.commit()
        self.rows.pop()
        self.write_workbook()
        self.run_import()
        with closing(sqlite3.connect(self.db)) as connection:
            self.assertEqual(connection.execute("SELECT source_id,count(*) FROM ruling_case GROUP BY source_id ORDER BY source_id").fetchall(),
                             [("cn_rulings_2023_2025", 1), ("other", 1)])
            self.assertEqual(connection.execute("SELECT product_name FROM ruling_case_fts WHERE source_id='other'").fetchone()[0], "其他原文")
            self.assertEqual(connection.execute("SELECT count(*) FROM ruling_code_map WHERE case_id='other:case'").fetchone()[0], 1)

    def test_insert_failure_rolls_back_deleted_source_cases_maps_and_fts(self):
        self.run_import()
        with closing(sqlite3.connect(self.db)) as connection:
            connection.execute("""CREATE TRIGGER abort_ruling BEFORE INSERT ON ruling_case BEGIN
                SELECT RAISE(ABORT, 'forced rollback'); END""")
            connection.commit()
        before = self.dump_database()
        self.rows[0]["product_name_cn"] = "不应提交的变化"
        self.write_workbook()
        with self.assertRaisesRegex(sqlite3.IntegrityError, "forced rollback"):
            self.run_import()
        self.assertEqual(self.dump_database(), before)

    def test_failed_first_import_rolls_back_all_new_ddl(self):
        with closing(sqlite3.connect(self.db)) as connection:
            connection.execute("CREATE TABLE ruling_case_fts(wrong_column TEXT)")
            connection.commit()
        before = self.dump_database()
        with self.assertRaises(sqlite3.OperationalError):
            self.run_import()
        self.assertEqual(self.dump_database(), before)

    def test_non_ruling_content_mutation_is_detected_and_rolled_back(self):
        self.run_import()
        with closing(sqlite3.connect(self.db)) as connection:
            connection.execute("""CREATE TRIGGER corrupt_rule AFTER INSERT ON ruling_case BEGIN
                UPDATE legal_rule SET full_text='changed' WHERE rule_id='rule:1'; END""")
            connection.commit()
        before = self.dump_database()
        with self.assertRaisesRegex(RuntimeError, "non-ruling|原有"):
            self.run_import()
        self.assertEqual(self.dump_database(), before)

    def test_reports_original_table_counts_and_content_hashes_unchanged(self):
        result = self.run_import()
        self.assertEqual(result["non_ruling_before"], result["non_ruling_after"])
        self.assertEqual(result["non_ruling_before"]["hs_code"]["count"], 3)
        self.assertEqual(result["non_ruling_before"]["legal_rule"]["count"], 1)
        self.assertIn("legal_clause_fts_data", result["non_ruling_before"])
        self.assertEqual(len(result["non_ruling_before"]["legal_rule"]["sha256"]), 64)

    def test_cross_source_case_id_collision_never_replaces_the_other_source(self):
        self.run_import()
        with closing(sqlite3.connect(self.db)) as connection:
            connection.execute("UPDATE ruling_source SET source_id='other'")
            connection.execute("UPDATE ruling_case SET source_id='other'")
            connection.execute("UPDATE ruling_case_fts SET source_id='other'")
            connection.commit()
        before = self.dump_database()
        with self.assertRaises(sqlite3.IntegrityError):
            self.run_import()
        self.assertEqual(self.dump_database(), before)

    def test_missing_prefix_is_reported_without_inventing_a_mapping(self):
        self.rows[0].update(hs_code_original="9901.0000", heading_4="9901", subheading_6="990100")
        self.write_workbook()
        result = self.run_import()
        self.assertEqual(result["missing_prefix_count"], 1)
        with closing(sqlite3.connect(self.db)) as connection:
            self.assertEqual(connection.execute(
                "SELECT count(*) FROM ruling_code_map WHERE case_id='cn_ruling:J2023-0001'"
            ).fetchone()[0], 0)

    def test_formula_cells_are_refused_not_substituted_with_cached_values(self):
        self.rows[0]["product_description"] = '=CONCAT("not", "source text")'
        self.write_workbook()
        with self.assertRaisesRegex(ValueError, "formulas"):
            self.run_import(dry_run=True)

    def test_backup_failure_aborts_without_creating_ruling_schema(self):
        self.backups.write_text("a file is not a backup directory", encoding="utf-8")
        before = database_files(self.db)
        with self.assertRaises(OSError):
            self.run_import()
        self.assertEqual(database_files(self.db), before)

    def test_rejects_invalid_code_lengths_levels_and_hierarchy(self):
        corruptions = [
            ("hs_code_original", "87116"), ("hs_code_original", "8711600"),
            ("hs_code_original", "871160000"), ("hs_code_original", "8711600000"),
            ("hs_code_original", "8711abc6000"), ("code_level", "6位"),
            ("code_level", 8), ("heading_4", "1704"), ("subheading_6", "871161"),
            ("knowledge_status", "待审核"), ("year", "2023年"),
        ]
        before = database_files(self.db)
        for column, invalid in corruptions:
            with self.subTest(column=column, invalid=invalid):
                original = self.rows[0][column]
                self.rows[0][column] = invalid
                self.write_workbook()
                with self.assertRaisesRegex(ValueError, column):
                    self.run_import()
                self.rows[0][column] = original
                self.assertEqual(database_files(self.db), before)

    def test_rejects_duplicate_decision_numbers_and_unexpected_headers(self):
        self.rows.append(dict(self.rows[0]))
        self.write_workbook()
        with self.assertRaisesRegex(ValueError, "decision_no"):
            self.run_import()
        workbook = load_workbook(self.xlsx)
        workbook.active.cell(1, 1, "unexpected")
        workbook.save(self.xlsx)
        workbook.close()
        with self.assertRaisesRegex(ValueError, "column|字段|表头"):
            self.run_import()

    def test_strict_default_refuses_unapproved_statistics_before_any_backup(self):
        before = database_files(self.db)
        with self.assertRaisesRegex(ValueError, "strict|严格"):
            importer.import_ruling_knowledge(self.xlsx, self.db, backup_dir=self.backups)
        self.assertEqual(database_files(self.db), before)
        self.assertFalse(self.backups.exists())

    def test_full_portable_workbook_against_read_only_real_catalog(self):
        # Preflight uses a verified temporary file copy, never a project DB connection.
        db = ROOT / "hs_copilot.db"
        before = database_files(db)
        result = importer.import_ruling_knowledge(ROOT / "tools/data/rulings-104.xlsx", db, dry_run=True)
        self.assertEqual(database_files(db), before)
        self.assertEqual(result["ruling_case_count"], 104)
        self.assertEqual(result["approved_count"], 104)
        self.assertEqual(result["unique_prefix_count"], 27)
        self.assertEqual(result["multiple_prefix_count"], 77)
        self.assertEqual(result["missing_prefix_count"], 0)
        self.assertEqual(result["publish_date_missing_count"], 93)
        self.assertEqual(result["classification_decision_missing_count"], 24)

    def test_full_portable_workbook_formal_import_on_temporary_catalog_backup(self):
        source = ROOT / "hs_copilot.db"
        before = database_files(source)
        clone = self.root / "full-catalog.db"
        with importer._read_only_snapshot(source) as snapshot:
            with closing(sqlite3.connect(clone)) as target:
                snapshot.backup(target)
        xlsx = ROOT / "tools/data/rulings-104.xlsx"
        result = importer.import_ruling_knowledge(xlsx, clone, backup_dir=self.backups)
        self.assertEqual(result["ruling_case_count"], 104)
        self.assertEqual(result["non_ruling_before"], result["non_ruling_after"])
        self.assertEqual(database_files(source), before)
        with closing(sqlite3.connect(clone)) as connection:
            self.assertEqual(connection.execute("SELECT count(*) FROM ruling_case").fetchone()[0], 104)
            self.assertEqual(connection.execute("SELECT count(*) FROM ruling_case_fts").fetchone()[0], 104)
            self.assertEqual(connection.execute("SELECT count(*) FROM ruling_code_map").fetchone()[0], result["mapping_count"])
            self.assertEqual(connection.execute("SELECT count(*) FROM ruling_case WHERE publish_date IS NULL").fetchone()[0], 93)
            self.assertEqual(connection.execute("SELECT count(*) FROM ruling_case WHERE classification_decision IS NULL").fetchone()[0], 24)
            self.assertEqual(connection.execute("PRAGMA foreign_key_check").fetchall(), [])


if __name__ == "__main__":
    unittest.main()
