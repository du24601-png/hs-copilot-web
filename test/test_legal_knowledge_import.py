import sqlite3
import tempfile
import unittest
from contextlib import closing
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.import_legal_knowledge import (
    detect_effect,
    extract_relations,
    import_legal_knowledge,
    parse_workbook,
    split_clauses,
)


class LegalKnowledgeImportTest(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.xlsx = self.root / "fixture.xlsx"
        self.db = self.root / "fixture.db"
        self._make_workbook()
        self._make_database()

    def tearDown(self):
        self.tmp.cleanup()

    def _make_workbook(self):
        workbook = Workbook()

        national = workbook.active
        national.title = "本国子目注释"
        national.append(["涉及税号（8位）", "涉及税号（点号）", "定义/对象", "本国子目注释（完整原文）"])
        national.append([
            "02089010",
            "0208.9010",
            "乳鸽肉",
            "本国子目0208.9010所称“乳鸽肉”，是指28日龄以内食用雏鸽的肉。",
        ])
        national.append([
            "13021100",
            "1302.1100",
            None,
            "子目1302.1100的鸦片，我国禁止进口。",
        ])

        gri = workbook.create_sheet("GRI归类总规则")
        gri.append(["序号", "规则", "规则原文", "PDF页码", "税则印刷页码"])
        gri.append([1, "规则一", "具有法律效力的归类，应按税目条文和有关类注或章注确定。", 16, 6])

        sections = workbook.create_sheet("类注")
        sections.append(["类序号", "类", "类名称", "包含章节", "状态", "类注全文", "PDF页码", "税则印刷页码"])
        sections.append([1, "第一类", "活动物；动物产品", "第1章-第2章", "有类注", "一、本类包括幼仔。\n二、干产品包括冷冻干燥产品。", 17, 7])
        sections.append([2, "第二类", "植物产品", "第3章", "无单独类注", None, None, None])

        chapters = workbook.create_sheet("章注")
        chapters.append(["章序号", "章", "所属类", "章名称", "状态", "章注全文", "PDF页码", "税则印刷页码"])
        chapters.append([1, "第一章", "第1类", "活动物", "有章注", "本章不包括税目30.02的产品。", 17, 7])
        chapters.append([2, "第二章", "第1类", "肉及食用杂碎", "有章注", "一、本章包括食用肉。\n二、本章不包括不适合食用的产品。", 30, 20])
        chapters.append([3, "第三章", "第2类", "鱼", "无单独章注", None, None, None])

        workbook.save(self.xlsx)

    def _make_database(self):
        connection = sqlite3.connect(self.db)
        connection.executescript(
            """
            CREATE TABLE hs_code (
              code TEXT PRIMARY KEY,
              name TEXT NOT NULL,
              subheading8 TEXT,
              chapter TEXT
            );
            INSERT INTO hs_code VALUES ('0208901000', '乳鸽肉', '02089010', '02');
            INSERT INTO hs_code VALUES ('1302110000', '鸦片液汁及浸膏', '13021100', '13');
            CREATE TABLE legal_note (
              id INTEGER PRIMARY KEY,
              level TEXT NOT NULL,
              ref TEXT,
              seq INTEGER,
              title TEXT,
              content TEXT NOT NULL,
              version TEXT,
              source TEXT
            );
            CREATE VIRTUAL TABLE legal_note_fts USING fts5(
              level UNINDEXED, ref UNINDEXED, title, content
            );
            """
        )
        connection.commit()
        connection.close()

    def test_clause_split_and_machine_derived_effects_are_deterministic(self):
        clauses = split_clauses("一、本章包括甲。\n二、本章不包括税目30.02的产品。", "chapter_note")
        self.assertEqual([item["label"] for item in clauses], ["一", "二"])
        self.assertEqual(detect_effect(clauses[1]["text"]), "excludes")
        self.assertTrue(any(
            relation["relation_type"] == "excludes"
            and relation["target_type"] == "heading"
            and relation["target_ref"] == "3002"
            for relation in extract_relations(clauses[1]["text"])
        ))

    def test_parse_workbook_separates_classification_rules_and_compliance_notice(self):
        parsed = parse_workbook(self.xlsx)
        self.assertEqual(len(parsed["rules"]), 6)
        self.assertEqual(sum(rule["decision_eligible"] for rule in parsed["rules"]), 5)
        notices = [rule for rule in parsed["rules"] if rule["rule_type"] == "compliance_notice"]
        self.assertEqual(len(notices), 1)
        self.assertEqual(notices[0]["scope_refs"], ["13021100"])
        self.assertEqual(len(parsed["sections"]), 2)
        self.assertEqual(len(parsed["chapters"]), 3)

    def test_import_is_transactional_idempotent_and_removes_only_empty_prototype(self):
        first = import_legal_knowledge(
            self.xlsx,
            self.db,
            backup_enabled=False,
            strict=False,
        )
        second = import_legal_knowledge(
            self.xlsx,
            self.db,
            backup_enabled=False,
            strict=False,
        )

        self.assertEqual(first["legal_rule_count"], 6)
        self.assertEqual(first["decision_rule_count"], 5)
        self.assertEqual(first["compliance_notice_count"], 1)
        self.assertEqual(first["unmapped_national_scope_count"], 0)
        self.assertEqual(second["legal_rule_count"], 6)

        with closing(sqlite3.connect(self.db)) as connection:
            tables = {row[0] for row in connection.execute("SELECT name FROM sqlite_master WHERE type='table'")}
            self.assertNotIn("legal_note", tables)
            self.assertIn("legal_rule", tables)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM legal_rule").fetchone()[0], 6)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM legal_source").fetchone()[0], 1)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM legal_rule_scope").fetchone()[0], 6)
            self.assertEqual(connection.execute("PRAGMA foreign_key_check").fetchall(), [])

    def test_nonempty_prototype_is_never_silently_dropped(self):
        connection = sqlite3.connect(self.db)
        connection.execute(
            "INSERT INTO legal_note(level, content) VALUES ('chapter', '已有数据')"
        )
        connection.commit()
        connection.close()

        with self.assertRaisesRegex(RuntimeError, "legal_note"):
            import_legal_knowledge(
                self.xlsx,
                self.db,
                backup_enabled=False,
                strict=False,
            )

    def test_failed_import_rolls_back_schema_and_prototype_migration(self):
        workbook = load_workbook(self.xlsx)
        workbook["GRI归类总规则"].append([1, "重复规则一", "重复主键用于触发回滚", 16, 6])
        workbook.save(self.xlsx)
        workbook.close()

        with self.assertRaises(sqlite3.IntegrityError):
            import_legal_knowledge(
                self.xlsx,
                self.db,
                backup_enabled=False,
                strict=False,
            )

        with closing(sqlite3.connect(self.db)) as connection:
            tables = {row[0] for row in connection.execute("SELECT name FROM sqlite_master WHERE type='table'")}
            self.assertIn("legal_note", tables)
            self.assertNotIn("legal_rule", tables)


if __name__ == "__main__":
    unittest.main()
